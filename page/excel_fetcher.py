def excel_data_fetcher(file_name, default_dir):
    import openpyxl
    import os
    file_name = os.path.join(default_dir, f'{file_name}.xlsx')
    wb = openpyxl.load_workbook(file_name)
    sheet_data = []
    for x in range(len(wb.sheetnames)):
        sheet = wb[wb.sheetnames[x]]
        single_list = []
        line = []
        for i in range(1, sheet.max_column + 1):
            for j in range(1, sheet.max_row + 1):
                single_list.append(sheet.cell(row=j, column=i).value)
            line.append(single_list)
            single_list = []
        sheet_data.append(line)

    result = []
    for m in sheet_data:
        out_dict = {'area': m[0][0],
                    'course': m[0][1],
                    'time': [str(y) for y in m[0][2:]],
                    'Monday': [str(y) for y in m[1][2:]],
                    'Tuesday': [str(y) for y in m[2][2:]],
                    'Wednesday': [str(y) for y in m[3][2:]],
                    'Thursday': [str(y) for y in m[4][2:]],
                    'Friday': [str(y) for y in m[5][2:]],
                    'Saturday': [str(y) for y in m[6][2:]],
                    'Sunday': [str(y) for y in m[7][2:]]}
        result.append(out_dict)

    return result
